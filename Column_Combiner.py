
# coding: utf-8
# In[68]:

import subprocess
import sys

def install(package):
    subprocess.call([sys.executable, "-m", "pip", "install", package])

try:
    import pandas as pd
    import numpy as np
    import xlrd

except:
    install('pandas')
    install('numpy')
    install('xlrd')



import pandas as pd
import numpy as np
import os


# In[69]:


if not os.path.exists("Combined"):
    os.mkdir("Combined") 


# In[70]:


excel_df_name = str(input("Please give the name of the table.\n"))
excel_df = pd.read_excel(excel_df_name)


# In[71]:


Col_Names_List = str(input("Please give the name of columns in order of priority. Use comma as a delimiter.\n"))
Col_List = Col_Names_List.split(",")


# In[72]:


excel_df = excel_df.fillna("")
excel_df.loc[:,"Result"] = ""


# In[73]:


for Col in Col_List:
    Empty_Result_Index = excel_df[excel_df['Result'] == ''].index
    print(Empty_Result_Index)
    excel_df.loc[Empty_Result_Index,"Result"] = excel_df.loc[Empty_Result_Index,Col]


# In[74]:


from datetime import datetime
timestamp= datetime.now().strftime("%Y%m%d-%H%M%S")
excel_df.to_excel("Combined\\Combined_"+timestamp +"_"+excel_df_name)


# In[75]:


"""

Без кавычек:

#%load file.py
%%writefile file.py  - в начале блока
%pycat  -
%run file.py
%lsmagic

from IPython.core.interactiveshell import InteractiveShell
InteractiveShell.ast_node_interactivity = "all"

===savetime===
from datetime import datetime
thetime= datetime.now().strftime("%Y%m%d-%H%M%S")

===openpyxl===
from openpyxl import load_workbook
from openpyxl import Workbook

wb = load_workbook()
wb_ws = wb.get_active_sheet()

wrwb = Workbook()
wrwb_ws = wrwb.get_active_sheet()

wb.save()

===numpy/pandas===
import pandas as pd
import numpy as np

excel_df = pd.read_excel()
csv_df = pd.read_csv()


df.to_excel()
df.to_csv()

writer = pd.ExcelWriter('',engine='xlsxwriter',options={})
df.to_excel(writer)
writer.save()


====requests/BeautifulSoup===
import requests
from bs4 import BeautifulSoup

page = requests.get("http://yandex.ru")
page.encoding = "windows-1251"
soup = BeautifulSoup(''.join(page.text), "html.parser\"),
soup.findAll("div")


===Files and directories===
import os
FileList = os.listdir()

#if not os.path.exists("Folder"):
#   os.mkdir("Folder") 

"""

